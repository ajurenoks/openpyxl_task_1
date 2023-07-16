from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
max_row=ws.max_row
total=0
for row in range(2,max_row+1):
    hour=ws['B' + str(row)].value
    rate=ws['C' + str(row)].value
    salary=hour*rate
    if (salary>3000):
        daudzums=daudzums+1
print(total)
wb.close()